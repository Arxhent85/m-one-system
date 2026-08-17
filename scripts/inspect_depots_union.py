import openpyxl
import json
import os

# 1. Parse DEPO MONE 2026.xlsx
def parse_depo(fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    sheet = wb['Tabelle1']
    items = {}
    for r in range(1, sheet.max_row + 1):
        sku_val = sheet.cell(r, 1).value
        name_val = sheet.cell(r, 2).value
        qty_val = sheet.cell(r, 5).value
        unit_val = sheet.cell(r, 6).value
        if sku_val is None:
            continue
        sku_clean = str(sku_val).replace('`', '').replace("'", "").strip()
        if not sku_clean or sku_clean.lower() == 'shifra':
            continue
        try:
            qty = int(round(float(qty_val))) if qty_val is not None else 0
        except:
            qty = 0
        items[sku_clean] = {
            'sku': sku_clean,
            'name': str(name_val or '').strip(),
            'stock': qty,
            'unit': str(unit_val or 'cope').strip()
        }
    return items

depo_mone = parse_depo('Aktuelle daten/DEPO MONE 2026.xlsx')
depo_mensuri = parse_depo('Aktuelle daten/MENSURI DEPO 2026.xlsx')
depo_qerimi = parse_depo('Aktuelle daten/DEPO QERIMI 2026.xlsx')

print(f"DEPO MONE: {len(depo_mone)} SKUs")
print(f"MENSURI: {len(depo_mensuri)} SKUs")
print(f"QERIMI: {len(depo_qerimi)} SKUs")

all_skus = sorted(set(list(depo_mone.keys()) + list(depo_mensuri.keys()) + list(depo_qerimi.keys())), key=lambda x: (int(x) if x.isdigit() else 999999, x))
print(f"Total Union of Depot SKUs: {len(all_skus)}")

# Check against existing stockStore.ts
for sku in all_skus:
    mone_qty = depo_mone.get(sku, {}).get('stock', 0)
    men_qty = depo_mensuri.get(sku, {}).get('stock', 0)
    qer_qty = depo_qerimi.get(sku, {}).get('stock', 0)
    name = (depo_mone.get(sku) or depo_mensuri.get(sku) or depo_qerimi.get(sku))['name']
    print(f"SKU {sku:6s} | MONE: {mone_qty:5d} | MENSURI: {men_qty:4d} | QERIMI: {qer_qty:4d} | Name: {name}")
