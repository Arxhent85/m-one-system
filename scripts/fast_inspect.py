import os
import openpyxl

folder_path = r'd:\M ONE SYSTEM APP\Aktuelle daten'
files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xlsm')]

out_file = r'd:\M ONE SYSTEM APP\scripts\excel_summary.txt'

with open(out_file, 'w', encoding='utf-8') as out:
    out.write("==================================================\n")
    out.write("ANALYSIS OF 'Aktuelle daten' EXCEL FILES\n")
    out.write("==================================================\n\n")

    for file_name in files:
        file_path = os.path.join(folder_path, file_name)
        out.write(f"FILE: {file_name}\n")
        try:
            # read_only=True is super fast
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            out.write(f"   Sheets ({len(wb.sheetnames)}): {wb.sheetnames}\n")
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                out.write(f"\n   --- Sheet: '{sheet_name}' ---\n")
                
                rows_printed = 0
                for row in sheet.iter_rows(values_only=True):
                    if any(row):
                        clean_row = [str(cell)[:40] if cell is not None else '' for cell in row[:12]]
                        out.write(f"      Row {rows_printed + 1}: {clean_row}\n")
                        rows_printed += 1
                        if rows_printed >= 8:
                            break
                if rows_printed == 0:
                    out.write("      [Sheet is empty]\n")
            wb.close()
        except Exception as e:
            out.write(f"   Error reading file: {e}\n")
        out.write("\n" + "-"*50 + "\n\n")

print("Done! Check excel_summary.txt")
