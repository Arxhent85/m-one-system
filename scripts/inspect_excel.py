import os
import openpyxl
import sys

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

folder_path = r'd:\M ONE SYSTEM APP\Aktuelle daten'
files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xlsm')]

print("==================================================")
print("ANALYSIS OF 'Aktuelle daten' EXCEL FILES")
print("==================================================\n")

for file_name in files:
    file_path = os.path.join(folder_path, file_name)
    print(f"FILE: {file_name}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"   Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n   --- Sheet: '{sheet_name}' (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
            
            rows_printed = 0
            for row in sheet.iter_rows(values_only=True):
                if any(row):  # Skip empty rows
                    clean_row = [str(cell)[:40] if cell is not None else '' for cell in row[:10]]
                    print(f"      Row {rows_printed + 1}: {clean_row}")
                    rows_printed += 1
                    if rows_printed >= 6:
                        break
            if rows_printed == 0:
                print("      [Sheet is empty]")
    except Exception as e:
        print(f"   Error reading file: {e}")
    print("\n" + "-"*50 + "\n")
