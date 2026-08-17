import openpyxl
import json
import os
from datetime import datetime, date

# 1. Load NEW DATA 2026.xlsx
wb = openpyxl.load_workbook('Aktuelle daten/NEW DATA 2026.xlsx', data_only=True)
sheet = wb['Tabelle1']
rows = list(sheet.iter_rows(values_only=True))[4:]

print(f"Loaded {len(rows)} data rows from 'NEW DATA 2026.xlsx'.")

# 2. Existing Customer metadata (to keep cities/locations/contact info intact)
existing_customers_map = {}
if os.path.exists('lib/mockCustomers.json'):
    with open('lib/mockCustomers.json', 'r', encoding='utf-8') as f:
        try:
            for c in json.load(f):
                cnum = str(c.get('customer_number') or '').strip()
                if cnum:
                    existing_customers_map[cnum] = c
        except Exception as e:
            print("Notice loading mockCustomers:", e)

# 3. Known SKU metadata (categories, prices)
sku_rates = {}
sku_names = {}

# Also load Provision.xlsx for reference
if os.path.exists('Aktuelle daten/Provision.xlsx'):
    wb_prov = openpyxl.load_workbook('Aktuelle daten/Provision.xlsx', data_only=True)
    sh_prov = wb_prov.active
    for r in list(sh_prov.iter_rows(values_only=True))[1:]:
        if r[0] is not None and r[1] is not None:
            s = str(int(r[0])) if isinstance(r[0], (int, float)) else str(r[0]).strip()
            sku_rates[s] = round(float(r[1]), 4)

# 4. Helper for driver assignment
def get_driver_and_location(cust_raw, agent_raw):
    c = str(cust_raw or '').strip().replace('-', '').replace(' ', '').replace('.', '')
    if c.startswith('0'):
        c = c[1:]
    if c.startswith('1'):
        return 'Qerimi', 'Fahrzeug 2 (Depo Qerimi)'
    elif c.startswith('2'):
        return 'Mensuri', 'Fahrzeug 1 (Depo Mensuri)'
    ag = str(agent_raw or '').lower()
    if 'qerimi' in ag:
        return 'Qerimi', 'Fahrzeug 2 (Depo Qerimi)'
    if 'mensuri' in ag:
        return 'Mensuri', 'Fahrzeug 1 (Depo Mensuri)'
    return 'Zentrale', 'Zentrales Hauptlager (M-ONE)'

# 5. Parse Orders and Items
orders = []
current_order = None
all_customers = {}

for idx, r in enumerate(rows):
    if all(cell is None for cell in r):
        continue
    
    fak_ref = r[0]
    dt_val = r[1]
    val_amt = r[2]
    status_val = r[3]
    cust_nr = r[5]
    cust_name = r[6]
    sku = r[7]
    art_name = r[8]
    qty = r[9]
    agent = r[10]
    rate = r[11]
    
    # Store SKU metadata
    clean_sku = str(sku).strip() if sku is not None else ''
    if clean_sku:
        if art_name:
            sku_names[clean_sku] = str(art_name).strip()
        if rate is not None and float(rate) > 0:
            sku_rates[clean_sku] = round(float(rate), 4)
            
    # New order trigger: val_amt is not None (in this Excel format, the first row of each invoice contains the total invoice amount in Col 2)
    is_new = (val_amt is not None)
    
    if is_new or current_order is None:
        if current_order:
            orders.append(current_order)
            
        dt_str = '2026-01-01T10:00:00.000Z'
        if isinstance(dt_val, (datetime, date)):
            dt_str = dt_val.strftime('%Y-%m-%dT10:00:00.000Z')
        elif dt_val:
            dt_str = str(dt_val)
            
        clean_cust_nr = str(cust_nr or '').strip()
        clean_cust_name = str(cust_name or fak_ref or 'Kunde').strip()
        driver_name, loc_name = get_driver_and_location(clean_cust_nr, agent)
        
        current_order = {
            'id': f"ord-2026-{len(orders)+1:04d}",
            'order_number': f"FK-2026-{len(orders)+1:04d}",
            'invoice_ref': str(fak_ref or '').strip(),
            'date': dt_str[:10],
            'created_at': dt_str,
            'status': 'completed',
            'payment_status': 'paid',
            'customer_number': clean_cust_nr,
            'customer_name': clean_cust_name,
            'driver_name': driver_name,
            'vehicle_location_name': loc_name,
            'total_amount': round(float(val_amt or 0), 2),
            'items': []
        }
        
        # Track customer
        if clean_cust_nr and clean_cust_nr not in all_customers:
            existing = existing_customers_map.get(clean_cust_nr, {})
            all_customers[clean_cust_nr] = {
                'id': f"cust-{clean_cust_nr}",
                'customer_number': clean_cust_nr,
                'company_name': clean_cust_name,
                'city': existing.get('city') or 'KOSOVO',
                'agent': f"{driver_name} ({loc_name})" if driver_name in ('Mensuri', 'Qerimi') else 'Zentrale Hauptlager',
                'phone': existing.get('phone') or '+383 44 123 456',
                'is_active': True,
                'total_orders': 0,
                'total_volume': 0.0,
            }

    # Add item
    if current_order and clean_sku:
        qty_num = float(qty or 1)
        current_order['items'].append({
            'sku': clean_sku,
            'name': sku_names.get(clean_sku, str(art_name or 'Artikel')),
            'qty': qty_num,
            'unit_price': 0.0, # will calculate below
            'total': 0.0,
            'rate': sku_rates.get(clean_sku, 0.0)
        })

if current_order:
    orders.append(current_order)

print(f"✅ Parsed {len(orders)} Orders with total {sum(len(o['items']) for o in orders)} items.")

# 6. Calculate item unit prices and line totals proportionally so sum(items.total) == order.total_amount
for o in orders:
    total_amt = o['total_amount']
    item_count = len(o['items'])
    total_qty = sum(it['qty'] for it in o['items'])
    
    if total_amt > 0 and total_qty > 0:
        avg_price = total_amt / total_qty
        running_sum = 0.0
        for i, it in enumerate(o['items']):
            if i == item_count - 1:
                it['total'] = round(total_amt - running_sum, 2)
            else:
                it['total'] = round(it['qty'] * avg_price, 2)
                running_sum += it['total']
            it['unit_price'] = round(it['total'] / it['qty'], 2) if it['qty'] > 0 else 0.0
    else:
        for it in o['items']:
            it['unit_price'] = 0.0
            it['total'] = 0.0

# 7. Update Customer totals
for o in orders:
    cnum = o['customer_number']
    if cnum in all_customers:
        all_customers[cnum]['total_orders'] += 1
        all_customers[cnum]['total_volume'] = round(all_customers[cnum]['total_volume'] + o['total_amount'], 2)

# 8. Save updated files
with open('lib/commissionRates.json', 'w', encoding='utf-8') as f:
    json.dump(sku_rates, f, indent=2, ensure_ascii=False)
print(f"✅ Saved {len(sku_rates)} commission rates to lib/commissionRates.json")

with open('lib/mock2026Sales.json', 'w', encoding='utf-8') as f:
    json.dump(orders, f, indent=2, ensure_ascii=False)
print(f"✅ Saved {len(orders)} orders to lib/mock2026Sales.json")

customer_list = sorted(all_customers.values(), key=lambda c: c['customer_number'])
with open('lib/mockCustomers.json', 'w', encoding='utf-8') as f:
    json.dump(customer_list, f, indent=2, ensure_ascii=False)
print(f"✅ Saved {len(customer_list)} customers to lib/mockCustomers.json")
