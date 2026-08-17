import openpyxl
import datetime

wb = openpyxl.load_workbook('Aktuelle daten/NEW DATA 2026.xlsx', data_only=True)
sheet = wb['Tabelle1']

agents = set()
dates = []
total_rows = 0
skus = set()
total_pieces = 0
total_val = 0.0

for r in range(5, sheet.max_row + 1):
    faktura = sheet.cell(r, 1).value
    datum = sheet.cell(r, 2).value
    wert = sheet.cell(r, 3).value
    cust_num = sheet.cell(r, 6).value
    cust_name = sheet.cell(r, 7).value
    sku = sheet.cell(r, 8).value
    art_name = sheet.cell(r, 9).value
    qty = sheet.cell(r, 10).value
    agent = sheet.cell(r, 11).value
    
    if sku is None and faktura is None and cust_name is None:
        continue
    
    total_rows += 1
    if agent is not None:
        agents.add(str(agent).strip())
    if sku is not None:
        skus.add(str(sku).strip())
    if qty is not None:
        try:
            total_pieces += float(qty)
        except:
            pass
    if wert is not None:
        try:
            total_val += float(wert)
        except:
            pass
    if isinstance(datum, datetime.datetime):
        dates.append(datum.strftime('%Y-%m-%d'))
    elif datum is not None:
        dates.append(str(datum)[:10])

print(f"Total rows: {total_rows}")
print(f"Total pieces: {total_pieces}")
print(f"Total value from Wert column: {total_val:.2f} EUR")
print(f"Agents: {agents}")
print(f"Date range: min={min(dates) if dates else 'N/A'} max={max(dates) if dates else 'N/A'}")
print(f"Total distinct SKUs: {len(skus)}")
