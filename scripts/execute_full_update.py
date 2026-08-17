import openpyxl
import json
import os
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

print("=== STARTING FULL M-ONE DATA UPDATE ===")

# 1. PARSE DEPOT FILES (Lagerbestände)
def parse_depo(fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    sheet = wb['Tabelle1']
    items = {}
    for r in range(1, sheet.max_row + 1):
        sku_val = sheet.cell(r, 1).value
        name_val = sheet.cell(r, 2).value
        qty_val = sheet.cell(r, 5).value
        unit_val = sheet.cell(r, 6).value
        if sku_val is None:
            continue
        sku_clean = str(sku_val).replace('`', '').replace("'", "").strip()
        if not sku_clean or sku_clean.lower() == 'shifra':
            continue
        try:
            qty = int(round(float(qty_val))) if qty_val is not None else 0
        except:
            qty = 0
        items[sku_clean] = {
            'sku': sku_clean,
            'name': str(name_val or '').strip(),
            'stock': qty,
            'unit': str(unit_val or 'cope').strip()
        }
    return items

depo_mone = parse_depo('Aktuelle daten/DEPO MONE 2026.xlsx')
depo_mensuri = parse_depo('Aktuelle daten/MENSURI DEPO 2026.xlsx')
depo_qerimi = parse_depo('Aktuelle daten/DEPO QERIMI 2026.xlsx')

print(f"Lagerbestände geladen:")
print(f"  - DEPO MONE 2026: {len(depo_mone)} Artikel, Gesamtbestand: {sum(x['stock'] for x in depo_mone.values())} Stk.")
print(f"  - MENSURI DEPO 2026: {len(depo_mensuri)} Artikel, Gesamtbestand: {sum(x['stock'] for x in depo_mensuri.values())} Stk.")
print(f"  - DEPO QERIMI 2026: {len(depo_qerimi)} Artikel, Gesamtbestand: {sum(x['stock'] for x in depo_qerimi.values())} Stk.")

# Master prices & commission rates reference
sku_rates = {}
if os.path.exists('Aktuelle daten/Provision.xlsx'):
    wb_prov = openpyxl.load_workbook('Aktuelle daten/Provision.xlsx', data_only=True)
    sh_prov = wb_prov.active
    for r in list(sh_prov.iter_rows(values_only=True))[1:]:
        if r[0] is not None and r[1] is not None:
            s = str(int(r[0])) if isinstance(r[0], (int, float)) else str(r[0]).strip()
            sku_rates[s] = round(float(r[1]), 4)

with open('lib/commissionRates.json', 'w', encoding='utf-8') as f:
    json.dump(sku_rates, f, indent=2, ensure_ascii=False)
print(f"Saved {len(sku_rates)} commission rates to lib/commissionRates.json")

# 2. PARSE NEW DATA 2026.xlsx (Verlauf)
wb_sales = openpyxl.load_workbook('Aktuelle daten/NEW DATA 2026.xlsx', data_only=True)
sh_sales = wb_sales['Tabelle1']
sales_rows = list(sh_sales.iter_rows(values_only=True))[4:]

def get_driver_and_location(cust_raw, agent_raw):
    c = str(cust_raw or '').strip().replace('-', '').replace(' ', '').replace('.', '')
    if c.startswith('0'):
        c = c[1:]
    if c.startswith('1') and c != '10000':
        return 'Qerimi', 'Fahrzeug 2 (Depo Qerimi)'
    elif c.startswith('2') and c != '20000':
        return 'Mensuri', 'Fahrzeug 1 (Depo Mensuri)'
    # 3xxxx and 4xxxx and HQ go to M-ONE Zentrale
    return 'Zentrale', 'Zentrales Hauptlager (M-ONE)'

orders = []
current_order = None

for idx, r in enumerate(sales_rows):
    if all(cell is None for cell in r):
        continue
    
    fak_ref = r[0]
    dt_val = r[1]
    val_amt = r[2]
    status_val = r[3]
    cust_nr = r[5]
    cust_name = r[6]
    sku = r[7]
    art_name = r[8]
    qty = r[9]
    agent = r[10]
    
    clean_sku = str(sku).strip() if sku is not None else ''
    if clean_sku.startswith('`') or clean_sku.startswith("'"):
        clean_sku = clean_sku[1:]
    if clean_sku.endswith('.0'):
        clean_sku = clean_sku[:-2]
    clean_sku = clean_sku.strip()
    
    is_new = (val_amt is not None)
    
    if is_new or current_order is None:
        if current_order:
            orders.append(current_order)
            
        dt_str = '2026-01-01T10:00:00.000Z'
        if isinstance(dt_val, (datetime, date)):
            dt_str = dt_val.strftime('%Y-%m-%dT10:00:00.000Z')
        elif dt_val:
            dt_str = str(dt_val)
            
        clean_cust_nr = str(cust_nr or '').strip()
        if clean_cust_nr.endswith('.0'):
            clean_cust_nr = clean_cust_nr[:-2]
        clean_cust_name = str(cust_name or fak_ref or 'Kunde').strip()
        driver_name, loc_name = get_driver_and_location(clean_cust_nr, agent)
        
        current_order = {
            'id': f"ord-2026-{len(orders)+1:04d}",
            'order_number': f"FK-2026-{len(orders)+1:04d}",
            'invoice_ref': str(fak_ref or '').strip(),
            'date': dt_str[:10],
            'created_at': dt_str,
            'status': 'completed',
            'payment_status': 'paid',
            'customer_number': clean_cust_nr,
            'customer_name': clean_cust_name,
            'driver_name': driver_name,
            'vehicle_location_name': loc_name,
            'total_amount': round(float(val_amt or 0), 2),
            'items': []
        }

    if current_order and clean_sku:
        try:
            qty_num = float(qty or 1)
        except:
            qty_num = 1.0
            
        current_order['items'].append({
            'sku': clean_sku,
            'name': str(art_name or 'Artikel').strip(),
            'qty': qty_num,
            'unit_price': 0.0,
            'total': 0.0,
            'rate': sku_rates.get(clean_sku, 0.0)
        })

if current_order:
    orders.append(current_order)

print(f"Parsed {len(orders)} Orders from NEW DATA 2026 with total {sum(len(o['items']) for o in orders)} items.")

# Calculate line totals & unit prices proportionally
for o in orders:
    total_amt = o['total_amount']
    item_count = len(o['items'])
    total_qty = sum(it['qty'] for it in o['items'])
    
    if total_amt > 0 and total_qty > 0:
        avg_price = total_amt / total_qty
        running_sum = 0.0
        for i, it in enumerate(o['items']):
            if i == item_count - 1:
                it['total'] = round(total_amt - running_sum, 2)
            else:
                it['total'] = round(it['qty'] * avg_price, 2)
                running_sum += it['total']
            it['unit_price'] = round(it['total'] / it['qty'], 2) if it['qty'] > 0 else 0.0
    else:
        for it in o['items']:
            it['unit_price'] = 0.0
            it['total'] = 0.0

with open('lib/mock2026Sales.json', 'w', encoding='utf-8') as f:
    json.dump(orders, f, indent=2, ensure_ascii=False)
print(f"Saved {len(orders)} orders to lib/mock2026Sales.json")

# 3. UPDATE CUSTOMER STATS & AGENT ASSIGNMENT (ALL 3xxxx -> M-ONE ZENTRALE)
with open('lib/mockCustomers.json', 'r', encoding='utf-8') as f:
    customers = json.load(f)

# Reset stats
for c in customers:
    c['total_orders'] = 0
    c['total_volume'] = 0.0
    
    # Update agent mapping: Miloti -> M ONE Zentrale
    cnum = str(c.get('customer_number') or '').strip().replace('-', '').replace(' ', '')
    if cnum.startswith('0'):
        cnum = cnum[1:]
    if cnum.startswith('1') and cnum != '10000':
        c['agent'] = 'Qerimi (Fahrzeug 2 (Depo Qerimi))'
    elif cnum.startswith('2') and cnum != '20000':
        c['agent'] = 'Mensuri (Fahrzeug 1 (Depo Mensuri))'
    elif cnum.startswith('3') and cnum != '30000':
        c['agent'] = 'M-ONE Zentrale (Zentrales Hauptlager)'
    elif cnum.startswith('4'):
        c['agent'] = 'M-ONE Zentrale (B2B Partner)'
    else:
        c['agent'] = 'M-ONE Zentrale (Hauptlager)'

cust_map = {str(c['customer_number']).strip(): c for c in customers}

for o in orders:
    cnum = str(o.get('customer_number') or '').strip()
    if cnum in cust_map:
        cust_map[cnum]['total_orders'] += 1
        cust_map[cnum]['total_volume'] = round(cust_map[cnum]['total_volume'] + float(o.get('total_amount') or 0), 2)

with open('lib/mockCustomers.json', 'w', encoding='utf-8') as f:
    json.dump(customers, f, indent=2, ensure_ascii=False)
print(f"Updated {len(customers)} customers in lib/mockCustomers.json (3xxxx successfully mapped to M-ONE Zentrale)")

# Summary stats per agent
agent_counts = {}
for c in customers:
    ag = c.get('agent')
    agent_counts[ag] = agent_counts.get(ag, 0) + 1

print("\nKundenverteilung nach Agenten:")
for ag, cnt in sorted(agent_counts.items()):
    print(f"  {ag}: {cnt} Kunden")

