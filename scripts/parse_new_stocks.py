import openpyxl
import json

def parse_file(fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    sheet = wb['Tabelle1']
    items = []
    for r in range(1, sheet.max_row + 1):
        sku_val = sheet.cell(r, 1).value
        name_val = sheet.cell(r, 2).value
        qty_val = sheet.cell(r, 5).value
        unit_val = sheet.cell(r, 6).value
        
        if sku_val is None:
            continue
        sku_clean = str(sku_val or '').replace('`', '').replace("'", "").strip()
        if not sku_clean or sku_clean.lower() == 'shifra':
            continue
        try:
            qty = float(qty_val) if qty_val is not None else 0.0
        except:
            qty = 0.0
        items.append({
            'row': r,
            'sku': sku_clean,
            'name': str(name_val or '').strip(),
            'qty': qty,
            'unit': str(unit_val or 'cope').strip()
        })
    return items

mone_items = parse_file('Aktuelle daten/DEPO MONE 2026.xlsx')
mensuri_items = parse_file('Aktuelle daten/MENSURI DEPO 2026.xlsx')
qerimi_items = parse_file('Aktuelle daten/DEPO QERIMI 2026.xlsx')

print(f"DEPO MONE 2026: {len(mone_items)} items, total qty={sum(x['qty'] for x in mone_items)}")
for it in mone_items:
    print(f"  {it['sku']}: {it['name']} -> {it['qty']} {it['unit']}")

print(f"\nMENSURI DEPO 2026: {len(mensuri_items)} items, total qty={sum(x['qty'] for x in mensuri_items)}")
for it in mensuri_items:
    print(f"  {it['sku']}: {it['name']} -> {it['qty']} {it['unit']}")

print(f"\nDEPO QERIMI 2026: {len(qerimi_items)} items, total qty={sum(x['qty'] for x in qerimi_items)}")
for it in qerimi_items:
    print(f"  {it['sku']}: {it['name']} -> {it['qty']} {it['unit']}")
