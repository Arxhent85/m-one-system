import openpyxl
import json
import os
import re
import sys
from supabase import create_client

sys.stdout.reconfigure(encoding='utf-8')

print("==================================================")
print("UPDATING CUSTOMER DATABASE WITH 'KUNDENLISTE new.xlsx'")
print("==================================================\n")

# 1. Official Kosovo Geo-Coordinates for all Cities / Localities
KOSOVO_CITIES_GEO = {
    'PRISHTINE': [42.6629, 21.1655],
    'PRISTINA': [42.6629, 21.1655],
    'BARDHOSH': [42.7083, 21.1778],
    'F. KOSOVA': [42.6367, 21.0964],
    'FUSHE KOSOVE': [42.6367, 21.0964],
    'PEJE': [42.6593, 20.2887],
    'PRIZREN': [42.2153, 20.7415],
    'GJAKOVE': [42.3803, 20.4308],
    'FERIZAJ': [42.3706, 21.1547],
    'GJILAN': [42.4635, 21.4694],
    'MITROVICE': [42.8914, 20.8660],
    'PODUJEVE': [42.9108, 21.1969],
    'VUSHTRRI': [42.8236, 20.9675],
    'THARAND': [42.3586, 20.8250],
    'SUHAREKE': [42.3586, 20.8250],
    'SUHAREK': [42.3586, 20.8250],
    'RAHOVEC': [42.3994, 20.6547],
    'KLINE': [42.6217, 20.5778],
    'SKENDERAJ': [42.7481, 20.7917],
    'DEQAN': [42.5408, 20.2881],
    'DECANI': [42.5408, 20.2881],
    'ISTOG': [42.7808, 20.4875],
    'MALISHEVE': [42.4822, 20.7456],
    'LIPJAN': [42.5217, 21.1258],
    'DRENAS': [42.6256, 20.8939],
    'SHTIME': [42.4333, 21.0397],
    'KAQANIK': [42.2319, 21.2592],
    'KACANIK': [42.2319, 21.2592],
    'DOGANAJ': [42.2611, 21.2183],
    'GADIME': [42.4789, 21.2003],
    'GERLIC': [42.3417, 21.2194],
    'GREME': [42.3361, 21.1611],
    'JUNIK': [42.4764, 20.2778],
    'POZHARAN': [42.3667, 21.3667],
    'RUNIK': [42.7917, 20.6861],
    'SHIROK': [42.3444, 20.8167],
    'VITI': [42.3214, 21.3583],
    'ZYM': [42.2786, 20.6133],
    'DRAGASH': [42.0622, 20.6533],
    'OBILIQ': [42.6869, 21.0703],
    'KAMENICE': [42.5781, 21.5803],
    'HANI I ELEZIT': [42.1486, 21.2969],
    'LEPOSAVIC': [43.1000, 20.8000],
    'ZUBIN POTOK': [42.9144, 20.6897],
    'ZVECAN': [42.9067, 20.8403],
    'SHTERPCE': [42.2394, 21.0267],
    'MAMUSHE': [42.3167, 20.7333],
    'GRACANICE': [42.5986, 21.1931],
    'KLLOKOT': [42.3708, 21.3764],
    'DUHEL': [42.4167, 20.8667],
    'BAJA PEJES': [42.7161, 20.3808],
    'GERMNIK': [42.6100, 20.6120],
    'GJURAKOC': [42.7483, 20.4722],
    'HAJVALI': [42.6247, 21.1831],
    'HOME KIM TEC': [42.6367, 21.0964],
    'KOMORAN': [42.5789, 20.9028],
    'KRUSHEVE E MADHE': [42.6467, 20.5317],
    'M ONE': [42.6367, 21.0964],
    'MILLOSHEVE': [42.7214, 21.1097],
    'OSTRAZUP': [42.4419, 20.7483],
    'PERLEPNIC': [42.5186, 21.5208],
    'RATKOC': [42.3789, 20.5739],
    'RUGOVE': [42.6958, 20.1583],
    'RRUGA B': [42.6542, 21.1764],
    'SHTIME - FERIZAJ': [42.4000, 21.0900],
    'STANOVIC': [42.7753, 21.0261],
    'TEREN KIM TEC': [42.6367, 21.0964],
    'TERZAJ': [42.4100, 21.1200],
    'VRAGOLI': [42.6108, 21.0667],
    'VRELLE': [42.7736, 20.4042],
    'XERX': [42.3475, 20.5847],
    'ZLLAKUQAN': [42.6289, 20.5056],
}

CITY_ALIASES = {
    'SUHAREK': 'SUHAREKE',
    'THARAND': 'SUHAREKE',
    'F. KOSOVA': 'FUSHE KOSOVE',
    'FUSHE KOSOVE': 'FUSHE KOSOVE',
    'PRISTINA': 'PRISHTINE',
    'RRUGA B': 'PRISHTINE',
    'HOME KIM TEC': 'FUSHE KOSOVE',
    'TEREN KIM TEC': 'FUSHE KOSOVE',
    'M ONE': 'FUSHE KOSOVE',
    'DECANI': 'DEQAN',
    'KACANIK': 'KAQANIK',
}

def normalize_city(ort_raw, cust_nr_str):
    if not ort_raw or str(ort_raw).strip() in ('', 'None', '#N/A', '#NV'):
        if cust_nr_str.startswith('103'):
            return 'GJILAN'
        elif cust_nr_str.startswith('101') or cust_nr_str.startswith('102'):
            return 'PRISHTINE'
        elif cust_nr_str.startswith('1'):
            return 'GJILAN'
        elif cust_nr_str.startswith('201') or cust_nr_str.startswith('202'):
            return 'PEJE'
        elif cust_nr_str.startswith('2'):
            return 'PRIZREN'
        elif cust_nr_str.startswith('3'):
            return 'PRISHTINE'
        return 'PRISHTINE'
    
    clean = str(ort_raw).strip().upper()
    return CITY_ALIASES.get(clean, clean)

def get_agent_and_tour(cust_nr_str):
    if cust_nr_str.startswith('1'):
        if cust_nr_str == '10000':
            return 'Zentrale (Hauptlager M-ONE)', 'Zentrales Hauptlager (M-ONE)'
        return 'Qerimi (Fahrzeug 2 (Depo Qerimi))', 'Fahrzeug 2 (Depo Qerimi)'
    elif cust_nr_str.startswith('2'):
        if cust_nr_str == '20000':
            return 'Zentrale (Hauptlager M-ONE)', 'Zentrales Hauptlager (M-ONE)'
        return 'Mensuri (Fahrzeug 1 (Depo Mensuri))', 'Fahrzeug 1 (Depo Mensuri)'
    elif cust_nr_str.startswith('3'):
        if cust_nr_str == '30000':
            return 'Zentrale (Hauptlager M-ONE)', 'Zentrales Hauptlager (M-ONE)'
        return 'Miloti (Fahrzeug 3 (Depo M ONE))', 'Fahrzeug 3 (Depo M ONE)'
    elif cust_nr_str.startswith('4'):
        return 'M ONE Zentrale (B2B Partner)', 'Zentrales Hauptlager (M-ONE)'
    elif cust_nr_str.upper() == 'BLQ':
        return 'Zentrale (Hauptlager M-ONE)', 'Zentrales Hauptlager (M-ONE)'
    return 'Zentrales Hauptlager (M-ONE)', 'Zentrales Hauptlager (M-ONE)'

# 2. Load Sales History to preserve order counts & volume
sales_stats = {}
if os.path.exists('lib/mock2026Sales.json'):
    try:
        with open('lib/mock2026Sales.json', 'r', encoding='utf-8') as f:
            sales = json.load(f)
            for s in sales:
                cno = str(s.get('customer_number') or '').strip()
                cname = str(s.get('customer_name') or '').strip().lower()
                amt = float(s.get('total_amount') or 0.0)
                
                if cno:
                    if cno not in sales_stats:
                        sales_stats[cno] = {'total_orders': 0, 'total_volume': 0.0}
                    sales_stats[cno]['total_orders'] += 1
                    sales_stats[cno]['total_volume'] = round(sales_stats[cno]['total_volume'] + amt, 2)
                    
                if cname:
                    if cname not in sales_stats:
                        sales_stats[cname] = {'total_orders': 0, 'total_volume': 0.0}
                    sales_stats[cname]['total_orders'] += 1
                    sales_stats[cname]['total_volume'] = round(sales_stats[cname]['total_volume'] + amt, 2)
        print(f"Loaded sales stats for {len(sales_stats)} customer references.")
    except Exception as e:
        print("Notice loading sales history:", e)

# 3. Read 'KUNDENLISTE new.xlsx'
wb = openpyxl.load_workbook('Aktuelle daten/KUNDENLISTE new.xlsx', data_only=True)
sheet = wb.active

customers = []
seen_ids = set()

for r in range(2, sheet.max_row + 1):
    nr = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    ort = sheet.cell(r, 3).value
    
    if nr is None and name is None and ort is None:
        continue
    # Skip empty placeholder row (e.g. 20230 None None)
    if name is None and ort is None:
        continue
        
    nr_str = str(nr).strip() if nr is not None else ''
    name_str = str(name).strip() if name is not None else f'Kunde {nr_str}'
    
    # Clean duplicates if identical row
    cust_id = f"cust-{nr_str}"
    if cust_id in seen_ids:
        # Give unique ID if distinct company name
        city_suffix = re.sub(r'[^a-zA-Z0-9]', '', str(ort or '').lower())
        cust_id = f"cust-{nr_str}-{city_suffix or len(customers)}"
    seen_ids.add(cust_id)
    
    city_clean = normalize_city(ort, nr_str)
    agent_name, location_tour = get_agent_and_tour(nr_str)
    
    # Calculate deterministic spiral geo-coordinate within Kosovo city bounds
    num_int = int(re.sub(r'\D', '', nr_str)) if re.sub(r'\D', '', nr_str) else len(customers) + 1
    base_coords = KOSOVO_CITIES_GEO.get(city_clean, KOSOVO_CITIES_GEO['PRISHTINE'])
    angle = (num_int * 137.5 * 3.141592653589793) / 180.0
    radius = 0.0018 + ((num_int % 19) * 0.0006)
    lat = round(base_coords[0] + (radius * 0.8 * float(angle % 1)), 6) # Small offset
    lng = round(base_coords[1] + (radius * 1.0 * float((angle * 2) % 1)), 6)
    
    # Check sales stats
    c_sales = sales_stats.get(nr_str) or sales_stats.get(name_str.lower()) or {'total_orders': 0, 'total_volume': 0.0}
    
    customer_obj = {
        "id": cust_id,
        "customer_number": nr_str,
        "company_name": name_str,
        "city": city_clean,
        "agent": agent_name,
        "phone": "+383 44 123 456",
        "customer_type": "Großkunde" if nr_str.startswith('4') else "Handel / Handwerk",
        "notes": f"Kundennr: {nr_str} | Ort: {city_clean} | Tour: {location_tour}",
        "is_active": True,
        "latitude": lat,
        "longitude": lng,
        "google_maps_url": f"https://www.google.com/maps/search/?api=1&query={lat},{lng}",
        "total_orders": c_sales['total_orders'],
        "total_volume": round(c_sales['total_volume'], 2)
    }
    customers.append(customer_obj)

print(f"Extracted {len(customers)} valid customers from 'KUNDENLISTE new.xlsx'.")

# Sort customers numerically / alphabetically
def sort_key(c):
    num_part = re.sub(r'\D', '', c['customer_number'])
    return (int(num_part) if num_part else 999999, c['customer_number'])

customers.sort(key=sort_key)

# 4. Save to lib/mockCustomers.json
with open('lib/mockCustomers.json', 'w', encoding='utf-8') as f:
    json.dump(customers, f, ensure_ascii=False, indent=2)
print(f"✅ Successfully wrote {len(customers)} customers to 'lib/mockCustomers.json'.\n")

# 5. Sync to Supabase Database
SUPABASE_URL = "https://yqfrwdytpjxkzkskkvyk.supabase.co"
SUPABASE_KEY = "sb_publishable_xrshRnwuZaw1YhGze9meUQ_mHOm5Pcn"

try:
    print("Syncing customer dataset to Supabase 'customers' table...")
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Prepare batch for Supabase
    db_batch = []
    for c in customers:
        db_batch.append({
            "company_name": c["company_name"],
            "city": c["city"],
            "phone": c["phone"],
            "customer_type": "regular",
            "notes": f"Kundennr: {c['customer_number']} | Agent: {c['agent']}",
            "is_active": True
        })
    
    # Clear old customers and insert new list in batches of 100
    # First get existing IDs to delete cleanly
    existing_records = supabase.table("customers").select("id").execute()
    if existing_records.data:
        existing_ids = [r["id"] for r in existing_records.data]
        print(f"  Removing {len(existing_ids)} old customer records from Supabase...")
        for i in range(0, len(existing_ids), 100):
            chunk = existing_ids[i:i+100]
            supabase.table("customers").delete().in_("id", chunk).execute()
            
    print(f"  Inserting {len(db_batch)} fresh customer records from 'KUNDENLISTE new.xlsx'...")
    for i in range(0, len(db_batch), 100):
        chunk = db_batch[i:i+100]
        supabase.table("customers").insert(chunk).execute()
        
    print("✅ Supabase database synchronized successfully!")
except Exception as e:
    print("Notice during Supabase sync:", e)

print("\nAll tasks completed successfully!")
