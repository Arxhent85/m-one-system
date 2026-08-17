import openpyxl
import datetime
import json

wb = openpyxl.load_workbook('Aktuelle daten/NEW DATA 2026.xlsx', data_only=True)
sheet = wb['Tabelle1']

orders = []
current_order = None

for r in range(5, sheet.max_row + 1):
    faktura_val = sheet.cell(r, 1).value
    datum_val = sheet.cell(r, 2).value
    wert_val = sheet.cell(r, 3).value
    cust_num_val = sheet.cell(r, 6).value
    cust_name_val = sheet.cell(r, 7).value
    sku_val = sheet.cell(r, 8).value
    art_name_val = sheet.cell(r, 9).value
    qty_val = sheet.cell(r, 10).value
    agent_val = sheet.cell(r, 11).value
    
    if sku_val is None and faktura_val is None and cust_name_val is None:
        continue
    
    sku_str = str(sku_val or '').replace('`', '').replace("'", "").strip()
    art_name = str(art_name_val or '').strip()
    try:
        qty = float(qty_val) if qty_val is not None else 1.0
    except:
        qty = 1.0
        
    date_str = ''
    if isinstance(datum_val, datetime.datetime):
        date_str = datum_val.strftime('%Y-%m-%d')
    elif datum_val is not None:
        date_str = str(datum_val)[:10]
        
    cust_num_clean = str(cust_num_val or '').strip()
    cust_name_clean = str(cust_name_val or '').strip()
    agent_clean = str(agent_val or '').strip()
    
    # If date and cust_name exist or wert exists, it could be a new order or line item
    # Let's see how orders are grouped in existing mock2026Sales.json vs NEW DATA 2026.xlsx
    
print("First 20 rows:")
for r in range(5, 25):
    row_vals = [sheet.cell(r, c).value for c in range(1, 12)]
    print(f"Row {r}: {row_vals}")
