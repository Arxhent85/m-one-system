import json
import openpyxl
import os

with open('lib/mockCustomers.json', 'r', encoding='utf-8') as f:
    custs = json.load(f)

existing_nos = {c.get('customer_number'): c for c in custs if c.get('customer_number')}
existing_names = {c.get('company_name', '').strip().lower(): c for c in custs}

wb = openpyxl.load_workbook('Aktuelle daten/2026 Sells.xlsx', data_only=True)
sheet = wb.active

new_custs = {}
for r in range(5, sheet.max_row + 1):
    faktura = sheet.cell(row=r, column=1).value
    cust_no = sheet.cell(row=r, column=5).value
    cust_name = sheet.cell(row=r, column=6).value
    agent = sheet.cell(row=r, column=10).value

    if cust_no is not None and str(cust_no).strip() not in ['#N/A', '#NV', 'None']:
        no_str = str(cust_no).strip()
        name_str = str(cust_name).strip() if cust_name and str(cust_name) not in ['#N/A', '#NV', 'None'] else (str(faktura).strip() if faktura else 'Laufkunde')
        
        if no_str not in existing_nos and name_str.lower() not in existing_names:
            agent_str = str(agent).strip() if agent else ('Qerimi' if no_str.startswith('1') else ('Mensuri' if no_str.startswith('2') else 'Zentrale'))
            
            # City estimation based on agent or default
            city = 'Prishtinë'
            lat, lng = 42.6629, 21.1655
            if 'qerimi' in agent_str.lower():
                city = 'Gjilan'
                lat, lng = 42.4635, 21.4694
            elif 'mensuri' in agent_str.lower():
                city = 'Prizren'
                lat, lng = 42.2153, 20.7415

            new_custs[no_str] = {
                "id": f"cust-2026-{no_str}",
                "customer_number": no_str,
                "company_name": name_str,
                "city": city,
                "phone": "+383 44 000 000",
                "customer_type": "Handel / Handwerk",
                "notes": f"Erfasst aus Verkaufsbuch 2026 (Tour: {agent_str})",
                "agent": agent_str,
                "is_active": True,
                "latitude": lat,
                "longitude": lng,
                "google_maps_url": f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
            }

added_count = 0
for no_str, c_obj in new_custs.items():
    custs.append(c_obj)
    added_count += 1

print(f"Added {added_count} missing customers to mockCustomers.json. Total now: {len(custs)}")

with open('lib/mockCustomers.json', 'w', encoding='utf-8') as f:
    json.dump(custs, f, ensure_ascii=False, indent=2)

print("Saved updated lib/mockCustomers.json successfully!")
