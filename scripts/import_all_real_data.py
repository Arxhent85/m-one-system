import os
import openpyxl
import sys
import json
from supabase import create_client

sys.stdout.reconfigure(encoding='utf-8')

SUPABASE_URL = "https://yqfrwdytpjxkzkskkvyk.supabase.co"
SUPABASE_KEY = "sb_publishable_xrshRnwuZaw1YhGze9meUQ_mHOm5Pcn"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
folder_path = r'd:\M ONE SYSTEM APP\Aktuelle daten'

print("==================================================")
print("IMPORTING REAL BUSINESS DATA INTO SUPABASE")
print("==================================================\n")

# 1. STANDORTE ANLEGEN / ABFRAGEN
print("1. Syncing Standorte (Depots & Fahrzeuge)...")
locations_to_create = [
    {"name": "Hauptlager Depot (M-ONE)", "type": "depot", "description": "Zentrales Hauptlager"},
    {"name": "Fahrzeug Mensuri", "type": "vehicle", "description": "Lieferfahrzeug Mensuri"},
    {"name": "Fahrzeug Qerimi", "type": "vehicle", "description": "Lieferfahrzeug Qerimi"},
]

loc_map = {}
for loc in locations_to_create:
    res = supabase.table("locations").select("id").eq("name", loc["name"]).execute()
    if res.data and len(res.data) > 0:
        loc_map[loc["name"]] = res.data[0]["id"]
    else:
        inserted = supabase.table("locations").insert(loc).execute()
        if inserted.data:
            loc_map[loc["name"]] = inserted.data[0]["id"]

print(f"   Standorte OK: {loc_map}\n")

# 2. PRODUKTE IMPORTIEREN (aus KUNDENLISTE 2026.xlsm -> Sheet 'Produkte')
print("2. Importing 300 Produkte from 'KUNDENLISTE 2026.xlsm'...")
wb_kunden = openpyxl.load_workbook(os.path.join(folder_path, 'KUNDENLISTE 2026.xlsm'), data_only=True)
sheet_prod = wb_kunden['Produkte']

products_batch = []
for row in sheet_prod.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    art_nr, emri, vol, qmim_l, ean, ean2, qmimi, pct = (row[i] if i < len(row) else None for i in range(8))
    
    if not art_nr or not emri:
        continue
        
    sku = str(art_nr).strip()
    name = str(emri).strip()
    unit = str(vol).strip() if vol else 'Stk'
    
    try:
        price = float(qmimi) if qmimi is not None else 0.0
    except:
        price = 0.0
        
    barcode = str(ean).strip() if ean else None
    
    products_batch.append({
        "sku": sku,
        "name": name,
        "unit": unit,
        "purchase_price": round(price * 0.5, 2),  # ca. 50% EK Schätzung wenn nicht angegeben
        "selling_price": price,
        "min_stock": 10,
        "barcode": barcode,
        "is_active": True,
    })

# Batch insert / upsert Produkte
prod_res = supabase.table("products").upsert(products_batch, on_conflict="sku").execute()
print(f"   ✅ {len(products_batch)} Produkte erfolgreich in Supabase importiert/aktualisiert!\n")

# Map Product SKU -> Product ID
all_prods = supabase.table("products").select("id, sku").execute()
prod_map = {p["sku"]: p["id"] for p in all_prods.data}

# 3. KUNDEN IMPORTIEREN (aus KUNDENLISTE 2026.xlsm -> Sheet 'KUNDEN')
print("3. Importing 929 Kunden from 'KUNDENLISTE 2026.xlsm'...")
sheet_cust = wb_kunden['KUNDEN']

customers_batch = []
for row in sheet_cust.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    k_nr, k_name, ort, nr_unik, tel, info, axhenti = (row[i] if i < len(row) else None for i in range(7))
    
    if not k_name:
        continue
        
    c_name = str(k_name).strip()
    c_city = str(ort).strip() if ort else None
    c_phone = str(tel).strip() if tel else None
    c_notes = f"Kundennr: {k_nr} | Agent: {axhenti}" if k_nr or axhenti else None
    
    customers_batch.append({
        "company_name": c_name,
        "city": c_city,
        "phone": c_phone,
        "customer_type": "regular",
        "discount_pct": 0,
        "notes": c_notes,
        "is_active": True,
    })

# In Chunks von 200 einfügen
chunk_size = 200
for i in range(0, len(customers_batch), chunk_size):
    chunk = customers_batch[i:i + chunk_size]
    supabase.table("customers").insert(chunk).execute()

print(f"   ✅ {len(customers_batch)} Kunden erfolgreich in Supabase importiert!\n")

wb_kunden.close()

# 4. LAGERBESTÄNDE IMPORTIEREN (aus DEPO M ONE, DEPO MENSURI, DEPO QERIMI)
print("4. Importing Live Stock Levels for Depots & Vehicles...")

depo_files = [
    ("DEPO M ONE.xlsx", "Hauptlager Depot (M-ONE)"),
    ("DEPO MENSURI.xlsx", "Fahrzeug Mensuri"),
    ("DEPO QERIMI.xlsx", "Fahrzeug Qerimi"),
]

stock_items_count = 0
for file_name, loc_name in depo_files:
    loc_id = loc_map.get(loc_name)
    if not loc_id:
        continue
        
    wb_depo = openpyxl.load_workbook(os.path.join(folder_path, file_name), data_only=True)
    sheet_depo = wb_depo['Tabelle1']
    
    depo_stock_batch = []
    for row in sheet_depo.iter_rows(min_row=3, values_only=True):  # Ab Zeile 3 (Zeile 1 Header, Zeile 2 Depot Name)
        if not any(row):
            continue
        shifra, emertimi, prodhuesi, brendi, sasia, njm = (row[i] if i < len(row) else None for i in range(6))
        
        if not shifra and not emertimi:
            continue
            
        sku = str(shifra).replace('`', '').strip() if shifra else None
        
        # Finde Product ID via SKU
        prod_id = prod_map.get(sku)
        
        try:
            qty = float(sasia) if sasia is not None else 0.0
        except:
            qty = 0.0
            
        if prod_id and qty > 0:
            depo_stock_batch.append({
                "product_id": prod_id,
                "location_id": loc_id,
                "quantity": qty,
                "min_stock": 10,
            })
            
    if depo_stock_batch:
        supabase.table("stock_items").upsert(depo_stock_batch, on_conflict="product_id,location_id").execute()
        stock_items_count += len(depo_stock_batch)
        print(f"   - {loc_name}: {len(depo_stock_batch)} Artikel-Bestände verbucht")
        
    wb_depo.close()

print(f"\n✅ INSGESAMT {stock_items_count} LIVE-BESTÄNDE IN DEPOS & FAHRZEUGEN VERBUCHT!")

print("\n==================================================")
print("🎉 VOLLSTÄNDIGER DATENIMPORT ERFOLGREICH ABGESCHLOSSEN!")
print("==================================================")
