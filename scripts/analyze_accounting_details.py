import openpyxl
import json
import os

wb = openpyxl.load_workbook('Aktuelle daten/2026 Sells.xlsx', data_only=True)
sheet = wb.active

skus = {}
customers = {}
agents = {}
monthly = {}

for r in range(5, sheet.max_row + 1):
    faktura = sheet.cell(row=r, column=1).value
    date_val = sheet.cell(row=r, column=2).value
    wert = sheet.cell(row=r, column=3).value
    cust_no = sheet.cell(row=r, column=5).value
    cust_name = sheet.cell(row=r, column=6).value
    sku = sheet.cell(row=r, column=7).value
    name = sheet.cell(row=r, column=8).value
    qty = sheet.cell(row=r, column=9).value
    agent = sheet.cell(row=r, column=10).value

    # Month
    if date_val:
        m_key = f"{date_val.year:04d}-{date_val.month:02d}"
        if m_key not in monthly:
            monthly[m_key] = {'wert': 0.0, 'invoices': 0, 'qty': 0}
        if wert is not None:
            monthly[m_key]['wert'] += float(wert)
            monthly[m_key]['invoices'] += 1
        if qty is not None:
            monthly[m_key]['qty'] += int(float(qty))

    # SKU
    if sku is not None:
        sku_str = str(sku).strip()
        name_str = str(name).strip() if name is not None else ''
        q = int(float(qty)) if qty is not None else 0
        if sku_str not in skus:
            skus[sku_str] = {'name': name_str, 'total_qty': 0, 'count': 0}
        skus[sku_str]['total_qty'] += q
        skus[sku_str]['count'] += 1

    # Customer
    c_name = str(cust_name).strip() if cust_name and str(cust_name) not in ['#N/A', '#NV', 'None'] else (str(faktura).strip() if faktura else 'Laufkunde')
    c_no = str(cust_no).strip() if cust_no and str(cust_no) not in ['#N/A', '#NV', 'None'] else '—'
    c_key = f"{c_no}_{c_name}"
    if c_key not in customers:
        customers[c_key] = {'no': c_no, 'name': c_name, 'wert': 0.0, 'invoices': 0, 'qty': 0}
    if wert is not None:
        customers[c_key]['wert'] += float(wert)
        customers[c_key]['invoices'] += 1
    if qty is not None:
        customers[c_key]['qty'] += int(float(qty))

    # Agent
    ag_str = str(agent).strip() if agent else 'Unbekannt'
    if ag_str not in agents:
        agents[ag_str] = {'wert': 0.0, 'invoices': 0, 'qty': 0}
    if wert is not None:
        agents[ag_str]['wert'] += float(wert)
        agents[ag_str]['invoices'] += 1
    if qty is not None:
        agents[ag_str]['qty'] += int(float(qty))

print("=== MONATLICHE BILANZ 2026 ===")
for m in sorted(monthly.keys()):
    d = monthly[m]
    print(f"Monat {m}: {d['invoices']:4d} Fakturen | {d['qty']:6d} Stück | {d['wert']:10.2f} €")

print("\n=== AGENTEN / TOUREN BILANZ ===")
for a, d in sorted(agents.items(), key=lambda x: x[1]['wert'], reverse=True):
    print(f"Agent {a:12s}: {d['invoices']:4d} Fakturen | {d['qty']:6d} Stück | {d['wert']:10.2f} €")

print("\n=== TOP 10 KUNDEN NACH UMSATZ ===")
top_custs = sorted(customers.values(), key=lambda x: x['wert'], reverse=True)
for c in top_custs[:10]:
    print(f"Kunde [{c['no']:6s}] {c['name'][:25]:25s}: {c['invoices']:3d} Fakturen | {c['qty']:5d} Stück | {c['wert']:9.2f} €")

print("\n=== TOP 10 PRODUKTE NACH STÜCKZAHL ===")
top_prods = sorted(skus.items(), key=lambda x: x[1]['total_qty'], reverse=True)
for sku, p in top_prods[:10]:
    print(f"SKU {sku:8s} | {p['name'][:30]:30s}: {p['total_qty']:6d} Stück in {p['count']:4d} Bestellungen")
