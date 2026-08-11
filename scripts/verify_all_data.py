import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
folder = r'd:\M ONE SYSTEM APP\Aktuelle daten'

# Build price map from KUNDENLISTE
wb_kunden = openpyxl.load_workbook(os.path.join(folder, 'KUNDENLISTE 2026.xlsm'), data_only=True, read_only=True)
s_prod = wb_kunden['Produkte']
price_map = {}
for row in s_prod.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    art_nr, emri, vol, qmim_l, ean, ean2, qmimi, pct = (row[i] if i < len(row) else None for i in range(8))
    if art_nr:
        sku = str(int(art_nr)) if isinstance(art_nr, float) else str(art_nr).strip()
        try: sell_price = float(qmimi) if qmimi else 0.0
        except: sell_price = 0.0
        try: buy_price = float(qmim_l) if qmim_l else round(sell_price*0.5, 2)
        except: buy_price = round(sell_price*0.5, 2)
        price_map[sku] = (buy_price, sell_price)

# Collect customers
sheet_cust = wb_kunden['KUNDEN']
cust_count = 0
for row in sheet_cust.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    k_nr, k_name = row[0], row[1]
    if k_name: cust_count += 1
wb_kunden.close()

print(f"KUNDENLISTE: {cust_count} Kunden, {len(price_map)} Preise geladen")

# Now read all 3 depots
depo_files = [
    ('DEPO M ONE.xlsx', 'M-ONE'),
    ('DEPO MENSURI.xlsx', 'MENSURI'),
    ('DEPO QERIMI.xlsx', 'QERIMI'),
]

all_products = {}  # sku -> (name, unit)
stock_data = {}    # (sku, depot) -> qty

for fname, label in depo_files:
    wb = openpyxl.load_workbook(os.path.join(folder, fname), data_only=True, read_only=True)
    s = wb['Tabelle1']
    count = 0
    for row in s.iter_rows(min_row=3, values_only=True):
        if not any(row): continue
        shifra, emertimi, prod, brand, sasia, njm = (row[j] if j < len(row) else None for j in range(6))
        if not shifra and not emertimi: continue
        sku = str(shifra).replace('`', '').strip() if shifra else ''
        if not sku: continue
        name = str(emertimi).strip() if emertimi else ''
        unit = str(njm).strip() if njm else 'cope'
        try: qty = float(sasia) if sasia is not None else 0.0
        except: qty = 0.0
        if name: all_products[sku] = (name, unit)
        stock_data[(sku, label)] = qty
        count += 1
    print(f"  {label}: {count} Positionen")
    wb.close()

print(f"\nGesamt unique Produkte: {len(all_products)}")
print("\nSKU | Name | Einkauf | Verkauf | M-ONE | MENSURI | QERIMI")
print("-"*100)
for sku in sorted(all_products.keys(), key=lambda x: int(x) if x.isdigit() else 0):
    name, unit = all_products[sku]
    bp, sp = price_map.get(sku, (0.0, 0.0))
    m1 = stock_data.get((sku, 'M-ONE'), 0)
    mn = stock_data.get((sku, 'MENSURI'), 0)
    qr = stock_data.get((sku, 'QERIMI'), 0)
    print(f"{sku} | {name[:40]} | {bp:.2f} | {sp:.2f} | {m1} | {mn} | {qr}")
