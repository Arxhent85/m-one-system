import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder_path = r'd:\M ONE SYSTEM APP\Aktuelle daten'
sql_out_path = r'd:\M ONE SYSTEM APP\supabase\seed_real_data.sql'

print("Generating CLEAN SQL seed file (with user_id NULLABLE fix)...")

sql_lines = [
    "-- ============================================================",
    "-- M ONE ERP — REAL BUSINESS SEED DATA (2026)",
    "-- Automatically generated from 'Aktuelle daten'",
    "-- ============================================================\n",
    "-- Fix: Allow NULL user_id for imported historical sales orders",
    "ALTER TABLE sales_orders ALTER COLUMN user_id DROP NOT NULL;\n",
]

# 1. STANDORTE
sql_lines.append("-- 1. STANDORTE (1 Hauptlager Depot + 2 Fahrzeug-Depots)")
sql_lines.append("""INSERT INTO locations (id, name, type, description) VALUES
  ('11111111-1111-1111-1111-111111111111', 'Hauptlager Depot (M-ONE)', 'depot', 'Zentrales Hauptlager'),
  ('22222222-2222-2222-2222-222222222222', 'Fahrzeug 1 (Depo Mensuri)', 'vehicle', 'Lieferfahrzeug Mensuri'),
  ('33333333-3333-3333-3333-333333333333', 'Fahrzeug 2 (Depo Qerimi)', 'vehicle', 'Lieferfahrzeug Qerimi')
ON CONFLICT (name) DO UPDATE SET type = EXCLUDED.type;\n""")

# 2. PRODUKTE (300 Artikelstamm)
sql_lines.append("-- 2. PRODUKTE (300 Artikelstamm)")
wb_kunden = openpyxl.load_workbook(os.path.join(folder_path, 'KUNDENLISTE 2026.xlsm'), data_only=True, read_only=True)
sheet_prod = wb_kunden['Produkte']

seen_skus = set()
prod_values = []

for row in sheet_prod.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    art_nr, emri, vol, qmim_l, ean, ean2, qmimi, pct = (row[i] if i < len(row) else None for i in range(8))
    
    if not art_nr or not emri:
        continue
        
    sku = str(art_nr).replace("'", "''").strip()
    if sku in seen_skus:
        continue
    seen_skus.add(sku)
    
    name = str(emri).replace("'", "''").strip()
    unit = str(vol).replace("'", "''").strip() if vol else 'Stk'
    
    try:
        price = float(qmimi) if qmimi is not None else 0.0
    except:
        price = 0.0
        
    purchase_price = round(price * 0.5, 2)
    barcode = f"'{str(ean).strip()}'" if ean else "NULL"
    
    prod_values.append(f"('{sku}', '{name}', '{unit}', {purchase_price}, {price}, 10, {barcode}, true)")

if prod_values:
    sql_lines.append("INSERT INTO products (sku, name, unit, purchase_price, selling_price, min_stock, barcode, is_active) VALUES")
    sql_lines.append(",\n".join(prod_values))
    sql_lines.append("ON CONFLICT (sku) DO UPDATE SET selling_price = EXCLUDED.selling_price, name = EXCLUDED.name;\n")

print(f"   - Generated {len(prod_values)} unique products.")

# 3. KUNDEN (798 Kundenkartei)
sql_lines.append("-- 3. KUNDEN (798 Kundenkartei)")
sheet_cust = wb_kunden['KUNDEN']

cust_values = []
for row in sheet_cust.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    k_nr, k_name, ort, nr_unik, tel, info, axhenti = (row[i] if i < len(row) else None for i in range(7))
    
    if not k_name:
        continue
        
    c_name = str(k_name).replace("'", "''").strip()
    ort_str = str(ort).replace("'", "''").strip() if ort else ""
    tel_str = str(tel).replace("'", "''").strip() if tel else ""
    
    c_city = f"'{ort_str}'" if ort_str else "NULL"
    c_phone = f"'{tel_str}'" if tel_str else "NULL"
    c_notes = f"'Kundennr: {k_nr} | Agent: {axhenti}'" if k_nr or axhenti else "NULL"
    
    cust_values.append(f"('{c_name}', {c_city}, {c_phone}, 'regular', 0, {c_notes}, true)")

if cust_values:
    for b in range(0, len(cust_values), 250):
        batch = cust_values[b:b+250]
        sql_lines.append("INSERT INTO customers (company_name, city, phone, customer_type, discount_pct, notes, is_active) VALUES")
        sql_lines.append(",\n".join(batch) + ";\n")

print(f"   - Generated {len(cust_values)} customers.")
wb_kunden.close()

# 4. LAGERBESTÄNDE
sql_lines.append("-- 4. LIVE-BESTÄNDE (Depot & Fahrzeuge)")

depo_config = [
    ("DEPO M ONE.xlsx", "11111111-1111-1111-1111-111111111111", "Hauptlager Depot"),
    ("DEPO MENSURI.xlsx", "22222222-2222-2222-2222-222222222222", "Fahrzeug Mensuri"),
    ("DEPO QERIMI.xlsx", "33333333-3333-3333-3333-333333333333", "Fahrzeug Qerimi"),
]

stock_totals = {}

for file_name, loc_id, loc_label in depo_config:
    wb_depo = openpyxl.load_workbook(os.path.join(folder_path, file_name), data_only=True, read_only=True)
    sheet_depo = wb_depo['Tabelle1']
    
    for row in sheet_depo.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        shifra, emertimi, prodhuesi, brendi, sasia, njm = (row[i] if i < len(row) else None for i in range(6))
        
        if not shifra and not emertimi:
            continue
            
        sku = str(shifra).replace('`', '').strip() if shifra else None
        if not sku or sku not in seen_skus:
            continue
            
        try:
            qty = float(sasia) if sasia is not None else 0.0
        except:
            qty = 0.0
            
        if qty > 0:
            key = (sku, loc_id)
            stock_totals[key] = stock_totals.get(key, 0.0) + qty
            
    wb_depo.close()

stock_values = [
    f"((SELECT id FROM products WHERE sku = '{sku}' LIMIT 1), '{loc_id}', {qty}, 10)"
    for (sku, loc_id), qty in stock_totals.items()
]

if stock_values:
    sql_lines.append("INSERT INTO stock_items (product_id, location_id, quantity, min_stock) VALUES")
    sql_lines.append(",\n".join(stock_values))
    sql_lines.append("ON CONFLICT (product_id, location_id) DO UPDATE SET quantity = EXCLUDED.quantity;\n")

print(f"   - Generated {len(stock_values)} deduplicated stock items.")

# 5. HISTORISCHE VERKÄUFE 2026
sql_lines.append("-- 5. HISTORISCHE VERKÄUFE 2026")
wb_sales = openpyxl.load_workbook(os.path.join(folder_path, 'Daten 2026.xlsx'), data_only=True, read_only=True)
sheet_sales = wb_sales['Sheet1']

sales_values = []
seen_order_nos = set()
row_idx = 0

for row in sheet_sales.iter_rows(min_row=4, values_only=True):
    if not any(row):
        continue
    faktura, datum, wert, status, a, k_nr, k_name, art_nr, art_name, stueck, agent, pct = (row[i] if i < len(row) else None for i in range(12))
    
    if not k_name or not wert:
        continue
        
    c_name = str(k_name).replace("'", "''").strip()
    if not c_name or c_name == '#N/A':
        continue
        
    try:
        val = float(wert)
    except:
        val = 0.0
        
    if val <= 0:
        continue
        
    row_idx += 1
    order_no = f"ORD-2026-{row_idx:04d}"
    if order_no in seen_order_nos:
        continue
    seen_order_nos.add(order_no)
    
    date_str = str(datum)[:10] if datum else '2026-01-01'
    agent_str = str(agent).strip() if agent else 'Unbekannt'
    loc_target = '22222222-2222-2222-2222-222222222222' if 'Mensuri' in agent_str else '33333333-3333-3333-3333-333333333333'
    
    sales_values.append(
        f"('{order_no}', (SELECT id FROM customers WHERE company_name = '{c_name}' LIMIT 1), '{loc_target}', 'confirmed', 'paid', {val}, {val}, '{date_str} 10:00:00')"
    )
    if len(sales_values) >= 300:
        break

if sales_values:
    sql_lines.append("INSERT INTO sales_orders (order_number, customer_id, location_id, status, payment_status, subtotal, total_amount, created_at) VALUES")
    sql_lines.append(",\n".join(sales_values))
    sql_lines.append("ON CONFLICT (order_number) DO NOTHING;\n")

print(f"   - Generated {len(sales_values)} clean sales orders.")
wb_sales.close()

with open(sql_out_path, 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_lines))

print(f"\nALL DATA GENERATED WITH user_id FIX!")
