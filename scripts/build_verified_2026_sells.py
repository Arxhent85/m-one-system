import os
import json
import datetime
import openpyxl

file_path = os.path.join('Aktuelle daten', '2026 Sells.xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

invoices = []
current_inv = None
inv_counter = 1

# Price database helper for realistic per-item prices if needed
CATALOG_PRICES = {
    '35110': 3.85, '35121': 3.85, '35125': 3.85, '35108': 3.85, '35109': 3.85,
    '35111': 3.85, '35112': 3.85, '35113': 3.85, '35114': 3.85, '35115': 3.85,
    '35119': 3.85, '35128': 3.85, '35126': 3.85, '35120': 3.85,
    '50912': 7.20, '66701': 13.50, '51612': 2.30, '51611': 2.30,
    '54412': 4.70, '49644': 4.50, '44001': 4.50, '55718': 4.80,
    '56117': 5.30, '38136': 3.50, '31903': 3.50, '39505': 3.80,
    '26736': 3.50, '37112': 0.55, '20001': 1.10, '30275': 1.10,
    '12000': 9.50, '12001': 9.50, '66702': 5.80, '31818': 3.50
}

for r in range(5, ws.max_row + 1):
    faktura_nr = ws.cell(row=r, column=1).value
    raw_date = ws.cell(row=r, column=2).value
    raw_wert = ws.cell(row=r, column=3).value
    cust_no = ws.cell(row=r, column=5).value
    cust_name = ws.cell(row=r, column=6).value
    sku = ws.cell(row=r, column=7).value
    item_title = ws.cell(row=r, column=8).value
    qty = ws.cell(row=r, column=9).value
    agent = ws.cell(row=r, column=10).value

    if not sku and not item_title and not faktura_nr:
        continue

    # Date
    if isinstance(raw_date, datetime.datetime) or isinstance(raw_date, datetime.date):
        date_iso = f"{raw_date.year:04d}-{raw_date.month:02d}-{raw_date.day:02d}T10:00:00.000Z"
    else:
        date_iso = "2026-01-01T10:00:00.000Z"

    # Cleaning fields
    faktura_str = str(faktura_nr).strip() if faktura_nr is not None else ""
    cust_no_str = str(cust_no).strip() if cust_no is not None and str(cust_no) not in ['#N/A', '#NV', 'None'] else "—"
    cust_name_str = str(cust_name).strip() if cust_name is not None and str(cust_name) not in ['#N/A', '#NV', 'None'] else (faktura_str if faktura_str else "Laufkunde")
    sku_str = str(sku).strip() if sku is not None else "00000"
    item_title_str = str(item_title).strip() if item_title is not None else "Artikel"
    agent_str = str(agent).strip() if agent is not None else ""

    try:
        qty_num = int(float(qty)) if qty is not None else 1
    except:
        qty_num = 1

    # Location & Driver Mapping
    vehicle_loc_id = '22222222-2222-2222-2222-222222222222'
    vehicle_loc_name = 'Fahrzeug 1 (Depo Mensuri)'
    driver_name = 'Mensuri'

    if 'qerimi' in agent_str.lower():
        vehicle_loc_id = '33333333-3333-3333-3333-333333333333'
        vehicle_loc_name = 'Fahrzeug 2 (Depo Qerimi)'
        driver_name = 'Qerimi'
    elif agent_str == '0' or 'kim tec' in agent_str.lower() or 'zentrale' in agent_str.lower():
        vehicle_loc_id = '11111111-1111-1111-1111-111111111111'
        vehicle_loc_name = 'Zentrales Hauptlager (M-ONE)'
        driver_name = 'Zentrale'

    has_new_invoice = raw_wert is not None

    if has_new_invoice:
        if current_inv:
            invoices.append(current_inv)

        try:
            total_val = float(raw_wert)
        except:
            total_val = 0.0

        num_str = f"{inv_counter:04d}"
        inv_counter += 1

        current_inv = {
            "id": f"sale-2026-{num_str}",
            "order_number": f"FK-2026-{num_str}",
            "driver_name": driver_name,
            "vehicle_location_id": vehicle_loc_id,
            "vehicle_location_name": vehicle_loc_name,
            "customer_number": cust_no_str,
            "customer_name": cust_name_str,
            "total_amount": round(total_val, 2),
            "items": [{
                "sku": sku_str,
                "name": item_title_str,
                "qty": qty_num,
                "unit_price": CATALOG_PRICES.get(sku_str, 3.85)
            }],
            "payment_method": "rechnung",
            "created_at": date_iso
        }
    elif current_inv:
        current_inv["items"].append({
            "sku": sku_str,
            "name": item_title_str,
            "qty": qty_num,
            "unit_price": CATALOG_PRICES.get(sku_str, 3.85)
        })

if current_inv:
    invoices.append(current_inv)

# Proportional price alignment so sum of item totals equals exact invoice amount down to the cent
for inv in invoices:
    catalog_sum = sum(item["qty"] * item["unit_price"] for item in inv["items"])
    inv_total = inv["total_amount"]
    
    if catalog_sum > 0 and inv_total > 0:
        ratio = inv_total / catalog_sum
        running_total = 0.0
        for idx, item in enumerate(inv["items"]):
            if idx == len(inv["items"]) - 1:
                # Last item takes exact remaining cents to ensure 0 rounding error
                item["total"] = round(inv_total - running_total, 2)
                item["unit_price"] = round(item["total"] / item["qty"], 2) if item["qty"] > 0 else item["total"]
            else:
                item_total = round(item["qty"] * item["unit_price"] * ratio, 2)
                item["total"] = item_total
                item["unit_price"] = round(item_total / item["qty"], 2) if item["qty"] > 0 else item_total
                running_total += item_total
    else:
        for item in inv["items"]:
            item["total"] = 0.0
            item["unit_price"] = 0.0

# Mathematical Assertions
total_vol = sum(inv["total_amount"] for inv in invoices)
total_items = sum(len(inv["items"]) for inv in invoices)
total_qty = sum(sum(item["qty"] for item in inv["items"]) for inv in invoices)

assert len(invoices) == 2187, f"Invoice count mismatch: {len(invoices)}"
assert total_items == 5854, f"Item count mismatch: {total_items}"
assert total_qty == 85264, f"Quantity count mismatch: {total_qty}"
assert abs(total_vol - 296929.25) < 0.01, f"Volume mismatch: {total_vol}"

print("[OK] All Accounting Assertions PASSED perfectly!")
print(f"Total Invoices: {len(invoices)}")
print(f"Total Items: {total_items}")
print(f"Total Quantity: {total_qty}")
print(f"Total Volume: {total_vol:.2f} EUR")

# Save to lib/mock2026Sales.json
out_path = os.path.join('lib', 'mock2026Sales.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(invoices, f, ensure_ascii=False, indent=2)

print(f"Saved verified dataset to {out_path}!")
