import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r'd:\M ONE SYSTEM APP\Aktuelle daten'
sql_out = r'd:\M ONE SYSTEM APP\supabase\seed_real_data.sql'

print("=== GENERATING FINAL EXACT SEED SQL (KUNDENNUMMER PREFIX RULES) ===")

# 1. PREISLISTE LADEN
wb_kunden = openpyxl.load_workbook(os.path.join(folder, 'KUNDENLISTE 2026.xlsm'), data_only=True, read_only=True)
s_prod = wb_kunden['Produkte']
price_map = {}
for row in s_prod.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    art_nr, emri, vol, qmim_l, ean, ean2, qmimi, pct = (row[i] if i < len(row) else None for i in range(8))
    if not art_nr: continue
    sku = str(int(art_nr)) if isinstance(art_nr, float) else str(art_nr).strip()
    try: sell_price = round(float(qmimi), 2) if qmimi else 0.0
    except: sell_price = 0.0
    try: buy_price = round(float(qmim_l), 2) if qmim_l else round(sell_price * 0.5, 2)
    except: buy_price = round(sell_price * 0.5, 2)
    price_map[sku] = (buy_price, sell_price)

# 2. PRODUKTE SAMMELN
all_products = {}
stock_data = {}

loc_map = {
    'M-ONE':   '11111111-1111-1111-1111-111111111111',
    'MENSURI': '22222222-2222-2222-2222-222222222222',
    'QERIMI':  '33333333-3333-3333-3333-333333333333',
}

depo_files = [
    ('DEPO M ONE.xlsx',    'M-ONE'),
    ('DEPO MENSURI.xlsx',  'MENSURI'),
    ('DEPO QERIMI.xlsx',   'QERIMI'),
]

for fname, label in depo_files:
    wb = openpyxl.load_workbook(os.path.join(folder, fname), data_only=True, read_only=True)
    s = wb['Tabelle1']
    for row in s.iter_rows(min_row=3, values_only=True):
        if not any(row): continue
        shifra, emertimi, prod, brand, sasia, njm = (row[j] if j < len(row) else None for j in range(6))
        if not shifra and not emertimi: continue
        sku = str(shifra).replace('`', '').strip() if shifra else ''
        if not sku: continue
        name = str(emertimi).strip() if emertimi else ''
        unit = str(njm).strip() if njm else 'cope'
        try: qty = float(sasia) if sasia is not None else 0.0
        except: qty = 0.0
        if name and sku not in all_products:
            all_products[sku] = (name, unit)
        stock_data[(sku, loc_map[label])] = qty
    wb.close()

# 3. KUNDEN MIT ECHTEN KUNDENNUMMERN & PRÄFIX-REGELN LADEN
sheet_cust = wb_kunden['KUNDEN']
cust_rows = []
cust_count = 0

for row in sheet_cust.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    k_nr, k_name, ort, nr_unik, tel, info, axhenti = (row[i] if i < len(row) else None for i in range(7))
    if not k_name: continue
    
    cust_count += 1
    c_nr_str = str(int(k_nr)) if isinstance(k_nr, (int, float)) else (str(k_nr).strip() if k_nr else f"10000")
    c_name = str(k_name).replace("'", "''").strip()
    ort_str = str(ort).replace("'", "''").strip() if ort else ''
    tel_str = str(tel).replace("'", "''").strip() if tel else ''
    
    # Prefix-Regel: 1xxxx = Qerimi, 2xxxx = Mensuri, 3xxxx = Miloti
    if c_nr_str.startswith('1'):
        agent_assigned = 'Qerimi'
    elif c_nr_str.startswith('2'):
        agent_assigned = 'Mensuri'
    elif c_nr_str.startswith('3'):
        agent_assigned = 'Miloti'
    else:
        agent_assigned = str(axhenti).replace("'", "''").strip() if axhenti else 'M ONE'
    
    c_city  = f"'{ort_str}'" if ort_str else 'NULL'
    c_phone = f"'{tel_str}'" if tel_str else 'NULL'
    c_notes = f"'Kundennr: {c_nr_str} | Agent: {agent_assigned}'"
    
    cust_rows.append(f"('{c_nr_str}', '{c_name}', {c_city}, {c_phone}, 'regular', 0, {c_notes}, true)")

wb_kunden.close()

# 4. SQL GENERIEREN
lines = [
    "-- ============================================================",
    "-- M ONE ERP — EXACT SEED DATA (EXACT 3 LOCATIONS & DRIVER RULES)",
    "-- 1xxxx = Qerimi | 2xxxx = Mensuri | 3xxxx = Miloti",
    "-- ============================================================\n",
    "ALTER TABLE customers ADD COLUMN IF NOT EXISTS customer_number VARCHAR(50);",
    "ALTER TABLE sales_orders ALTER COLUMN user_id DROP NOT NULL;\n",
    "DELETE FROM stock_items;",
    "DELETE FROM sales_orders;",
    "DELETE FROM products;",
    "DELETE FROM customers;",
    "DELETE FROM locations;\n",
]

# Standorte
lines.append("-- 1. STANDORTE (EXAKT 3 STANDORTE)")
lines.append("""INSERT INTO locations (id, name, type, description) VALUES
  ('11111111-1111-1111-1111-111111111111', 'Hauptlager Depot (M-ONE)', 'depot', 'Zentrales Hauptlager'),
  ('22222222-2222-2222-2222-222222222222', 'Fahrzeug 1 (Depo Mensuri)', 'vehicle', 'Lieferfahrzeug Mensuri'),
  ('33333333-3333-3333-3333-333333333333', 'Fahrzeug 2 (Depo Qerimi)', 'vehicle', 'Lieferfahrzeug Qerimi');
""")

# Produkte
lines.append("-- 2. PRODUKTE")
prod_vals = []
for sku in sorted(all_products.keys(), key=lambda x: int(x) if x.isdigit() else 0):
    name, unit = all_products[sku]
    name_esc = name.replace("'", "''")
    unit_esc = unit.replace("'", "''")
    bp, sp = price_map.get(sku, (0.0, 0.0))
    prod_vals.append(f"('{sku}', '{name_esc}', '{unit_esc}', {bp}, {sp}, 0, true)")

lines.append("INSERT INTO products (sku, name, unit, purchase_price, selling_price, min_stock, is_active) VALUES")
lines.append(",\n".join(prod_vals))
lines.append("ON CONFLICT (sku) DO UPDATE SET name = EXCLUDED.name, selling_price = EXCLUDED.selling_price;\n")

# Kunden mit präziser Kundennummer & Agent
lines.append("-- 3. KUNDEN MIT PRÄFIX-ZUTEILUNG (1xxxx=Qerimi, 2xxxx=Mensuri)")
for b in range(0, len(cust_rows), 200):
    batch = cust_rows[b:b+200]
    lines.append("INSERT INTO customers (customer_number, company_name, city, phone, customer_type, discount_pct, notes, is_active) VALUES")
    lines.append(",\n".join(batch) + " ON CONFLICT DO NOTHING;\n")

# Bestände
lines.append("-- 4. LAGERBESTÄNDE")
stock_vals = []
for (sku, loc_id), qty in sorted(stock_data.items(), key=lambda x: (int(x[0][0]) if x[0][0].isdigit() else 0, x[0][1])):
    if qty <= 0: continue
    stock_vals.append(f"((SELECT id FROM products WHERE sku='{sku}' LIMIT 1), '{loc_id}', {qty}, 0)")

lines.append("INSERT INTO stock_items (product_id, location_id, quantity, min_stock) VALUES")
lines.append(",\n".join(stock_vals))
lines.append("ON CONFLICT (product_id, location_id) DO UPDATE SET quantity = EXCLUDED.quantity;\n")

with open(sql_out, 'w', encoding='utf-8') as f:
    f.write("\n".join(lines))

print(f"SUCCESSFULLY GENERATED SEED SQL WITH 1xxxx / 2xxxx DRIVER PREFIX RULES!")
