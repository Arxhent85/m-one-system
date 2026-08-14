import os
import json
import datetime
import openpyxl

file_path = os.path.join('Aktuelle daten', '2026 Sells.xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

invoices = []
current_inv = None
inv_counter = 1

for r in range(5, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
    faktura_nr = row_vals[0]
    raw_date = row_vals[1]
    raw_wert = row_vals[2]
    cust_no = row_vals[4]
    cust_name = row_vals[5]
    sku = row_vals[6]
    item_title = row_vals[7]
    qty = row_vals[8]
    agent = row_vals[9]

    if not sku and not item_title and not faktura_nr:
        continue

    # Format date
    date_iso = "2026-01-01T10:00:00.000Z"
    if isinstance(raw_date, datetime.datetime) or isinstance(raw_date, datetime.date):
        date_iso = f"{raw_date.year:04d}-{raw_date.month:02d}-{raw_date.day:02d}T10:00:00.000Z"
    elif isinstance(raw_date, str) and '.' in raw_date:
        parts = raw_date.strip().split('.')
        if len(parts) == 3:
            try:
                # Could be M.D.YYYY or D.M.YYYY
                p0, p1, p2 = int(parts[0]), int(parts[1]), int(parts[2])
                if p0 > 12: # D.M.YYYY
                    date_iso = f"{p2:04d}-{p1:02d}-{p0:02d}T10:00:00.000Z"
                else: # M.D.YYYY
                    date_iso = f"{p2:04d}-{p0:02d}-{p1:02d}T10:00:00.000Z"
            except:
                pass

    # Clean strings
    faktura_str = str(faktura_nr).strip() if faktura_nr is not None else ""
    cust_no_str = str(cust_no).strip() if cust_no is not None and str(cust_no) not in ['#N/A', '#NV', 'None'] else "—"
    cust_name_str = str(cust_name).strip() if cust_name is not None and str(cust_name) not in ['#N/A', '#NV', 'None'] else faktura_str
    sku_str = str(sku).strip() if sku is not None else "00000"
    item_title_str = str(item_title).strip() if item_title is not None else "Artikel"
    agent_str = str(agent).strip() if agent is not None else ""

    try:
        qty_num = int(float(qty)) if qty is not None else 1
    except:
        qty_num = 1

    # Location & Driver mapping
    vehicle_loc_id = '22222222-2222-2222-2222-222222222222'
    vehicle_loc_name = 'Fahrzeug 1 (Depo Mensuri)'
    driver_name = 'Mensuri'

    if 'qerimi' in agent_str.lower():
        vehicle_loc_id = '33333333-3333-3333-3333-333333333333'
        vehicle_loc_name = 'Fahrzeug 2 (Depo Qerimi)'
        driver_name = 'Qerimi'
    elif agent_str == '0' or 'kim tec' in agent_str.lower() or 'zentrale' in agent_str.lower():
        vehicle_loc_id = '11111111-1111-1111-1111-111111111111'
        vehicle_loc_name = 'Zentrales Hauptlager (M-ONE)'
        driver_name = 'Zentrale'

    has_new_invoice = raw_wert is not None

    if has_new_invoice:
        if current_inv:
            invoices.append(current_inv)

        try:
            total_val = float(raw_wert)
        except:
            total_val = 0.0

        num_str = f"{inv_counter:04d}"
        inv_counter += 1

        current_inv = {
            "id": f"sale-2026-{num_str}",
            "order_number": f"FK-2026-{num_str}",
            "driver_name": driver_name,
            "vehicle_location_id": vehicle_loc_id,
            "vehicle_location_name": vehicle_loc_name,
            "customer_number": cust_no_str,
            "customer_name": cust_name_str,
            "total_amount": round(total_val, 2),
            "items": [{
                "sku": sku_str,
                "name": item_title_str,
                "qty": qty_num,
                "unit_price": 3.85
            }],
            "payment_method": "rechnung",
            "created_at": date_iso
        }
    elif current_inv:
        current_inv["items"].append({
            "sku": sku_str,
            "name": item_title_str,
            "qty": qty_num,
            "unit_price": 3.85
        })

if current_inv:
    invoices.append(current_inv)

# Compute item unit prices & line totals to match exact invoice totals
for inv in invoices:
    total_qty = sum(item["qty"] for item in inv["items"])
    avg_price = (inv["total_amount"] / total_qty) if (inv["total_amount"] > 0 and total_qty > 0) else 3.85
    for item in inv["items"]:
        item["unit_price"] = round(avg_price, 2)
        item["total"] = round(item["qty"] * item["unit_price"], 2)

total_vol = sum(inv["total_amount"] for inv in invoices)
total_items = sum(len(inv["items"]) for inv in invoices)
total_qty = sum(sum(item["qty"] for item in inv["items"]) for inv in invoices)

print(f"Total Invoices: {len(invoices)}")
print(f"Total Line Items: {total_items}")
print(f"Total Quantity (Stück): {total_qty}")
print(f"Total Volume: {total_vol:.2f} €")

out_path = os.path.join('lib', 'mock2026Sales.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(invoices, f, ensure_ascii=False, indent=2)

print(f"Successfully written to {out_path}!")
