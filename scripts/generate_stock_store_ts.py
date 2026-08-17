import openpyxl
import json
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Parse existing prices from stockStore.ts
with open('lib/stockStore.ts', 'r', encoding='utf-8') as f:
    text = f.read()

existing_prices = {}
for line in text.splitlines():
    if 'sku:' in line and 'selling_price:' in line:
        m_sku = re.search(r"sku:\s*['\"]([^'\"]+)['\"]", line)
        m_name = re.search(r"name:\s*['\"]([^'\"]+)['\"]", line)
        m_pur = re.search(r"purchase_price:\s*([0-9.]+)", line)
        m_sel = re.search(r"selling_price:\s*([0-9.]+)", line)
        if m_sku:
            sku = m_sku.group(1)
            existing_prices[sku] = {
                'name': m_name.group(1) if m_name else '',
                'purchase_price': float(m_pur.group(1)) if m_pur else 1.50,
                'selling_price': float(m_sel.group(1)) if m_sel else 3.00,
            }

def parse_depo(fname):
    wb = openpyxl.load_workbook(fname, data_only=True)
    sheet = wb['Tabelle1']
    items = {}
    for r in range(1, sheet.max_row + 1):
        sku_val = sheet.cell(r, 1).value
        name_val = sheet.cell(r, 2).value
        qty_val = sheet.cell(r, 5).value
        unit_val = sheet.cell(r, 6).value
        if sku_val is None:
            continue
        sku_clean = str(sku_val).replace('`', '').replace("'", "").strip()
        if not sku_clean or sku_clean.lower() == 'shifra':
            continue
        try:
            qty = int(round(float(qty_val))) if qty_val is not None else 0
        except:
            qty = 0
        items[sku_clean] = {
            'sku': sku_clean,
            'name': str(name_val or '').strip(),
            'stock': qty,
            'unit': str(unit_val or 'cope').strip()
        }
    return items

depo_mone = parse_depo('Aktuelle daten/DEPO MONE 2026.xlsx')
depo_mensuri = parse_depo('Aktuelle daten/MENSURI DEPO 2026.xlsx')
depo_qerimi = parse_depo('Aktuelle daten/DEPO QERIMI 2026.xlsx')

all_skus = sorted(set(list(depo_mone.keys()) + list(depo_mensuri.keys()) + list(depo_qerimi.keys())), key=lambda x: (int(x) if x.isdigit() else 999999, x))

print("=== INITIAL_DEPO_PRODUCTS ===")
for sku in all_skus:
    mone_stock = depo_mone.get(sku, {}).get('stock', 0)
    # Prefer existing clean German/official product names if available, else depot file name
    name = existing_prices.get(sku, {}).get('name') or depo_mone.get(sku, {}).get('name') or depo_mensuri.get(sku, {}).get('name') or depo_qerimi.get(sku, {}).get('name') or 'Artikel'
    pur = existing_prices.get(sku, {}).get('purchase_price', 1.50)
    sel = existing_prices.get(sku, {}).get('selling_price', 3.00)
    unit = depo_mone.get(sku, {}).get('unit', 'cope')
    print(f"  {{ id: 'p-{sku}', sku: '{sku}', name: '{name}', stock: {mone_stock}, unit: '{unit}', purchase_price: {pur:.2f}, selling_price: {sel:.2f} }},")

print("\n=== INITIAL_MENSURI_STOCK ===")
for sku in sorted(depo_mensuri.keys(), key=lambda x: (int(x) if x.isdigit() else 999999, x)):
    q = depo_mensuri[sku]['stock']
    print(f"  '{sku}': {q},")

print("\n=== INITIAL_QERIMI_STOCK ===")
for sku in sorted(depo_qerimi.keys(), key=lambda x: (int(x) if x.isdigit() else 999999, x)):
    q = depo_qerimi[sku]['stock']
    print(f"  '{sku}': {q},")
